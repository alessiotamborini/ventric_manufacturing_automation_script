"""
Docstring for cuffldt.py

This module defines the Cuff Leakage Detection Tool (CuffLDT) class. 
This tool is specifically designed for Ventric Health's cuff manufacturing
quality assurance pipeline. The objective of CuffLDT is to analyze batches
of cuff runs to detect faulty units during the manufacturing process. 
"""

import os
import sys
import numpy as np
import pandas as pd
import json
import re
from concurrent.futures import ProcessPoolExecutor, as_completed
from matplotlib import pyplot as plt
import matplotlib.style as mplstyle
from openpyxl import load_workbook
import subprocess

from datetime import datetime
from PyQt5.QtWidgets import QPlainTextEdit
from PyQt5.QtGui import QFont
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QObject

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QFileDialog, QMessageBox,
    QVBoxLayout, QHBoxLayout, QWidget, QPushButton, QLabel, QLineEdit,
    QProgressBar, QCheckBox
)

DEFAULT_GEOMETRY_PATH = '.'

# ============================================================================
# MODULE-LEVEL PLOT WORKERS  (must be at module level to be picklable)
# ============================================================================

def _init_plot_worker():
    """Called once in each worker process: switch to fast non-interactive backend."""
    plt.switch_backend('agg')
    mplstyle.use('fast')


def _plot_worker(args):
    """Render and save one signal visualization. Runs in a worker process."""
    sample_data, file_name, output_folder = args
    try:
        # Extract cuff data
        cuff_values = np.array(sample_data['tester_info']['cuff_data']['cuff_values'])
        time_values = np.array(sample_data['tester_info']['cuff_data']['time'])

        # Isolate sSBP hold segment
        ssbp_hold = sample_data['hold_ssbp_cuff']
        length = len(ssbp_hold)
        ssbp_hold_data = cuff_values[-length:]
        gap = len(ssbp_hold_data) - np.argmax(np.abs(np.diff(ssbp_hold_data)))
        ssbp_hold_data = cuff_values[-(length + gap):-gap]

        fig, ax = plt.subplots(1, 2, figsize=(14, 5),
                               gridspec_kw={'width_ratios': [2, 1]}, sharey=True)

        ax[0].plot(time_values, cuff_values)
        ax[0].grid(axis='y')
        ax[0].set_title(f'Full Cuff Signal - {file_name}')
        ax[0].set_xlabel('Time (n)')
        ax[0].set_ylabel('Signal Amplitude')

        ax[1].plot(ssbp_hold_data, label='Signal')
        ax[1].axhline(np.mean(ssbp_hold_data[-10000:]), color='k',
                      linestyle='-', label='Mean of Settled Signal')
        ax[1].grid(axis='y')
        ax[1].set_title('sSBP Hold Signal')
        ax[1].set_xlabel('Time (n)')
        ax[1].legend()
        ax[1].set_yticks(np.arange(0, 17000, 2000))

        plt.tight_layout()
        plot_path = os.path.join(output_folder, f"{os.path.splitext(file_name)[0]}.png")
        plt.savefig(plot_path, dpi=100)
        plt.close(fig)
        return True
    except Exception:
        return False


class AnalysisWorker(QObject):
    progress = pyqtSignal(str)
    progress_update = pyqtSignal(str, int, int)  # (step_label, current, total)
    error = pyqtSignal(str)
    finished = pyqtSignal()

    def __init__(self, controller, generate_plots=True):
        super().__init__()
        self.controller = controller  # reference to CuffLDT but worker must not touch widgets
        self.generate_plots = generate_plots

    def run(self):
        try:
            # ── Step 1: results folder ────────────────────────────────────────
            self.progress.emit("[1/4] Setting up results folder...")
            self.controller._create_results_folder()
            self.progress.emit(f"      > {self.controller.results_folder}")

            # ── Step 2: load JSON files ───────────────────────────────────────
            self.progress.emit("[2/4] Loading JSON files...")

            def load_cb(current, total):
                self.progress_update.emit("Loading files", current, total)

            data, files = self.controller._load_json_files(progress_callback=load_cb)
            if data is None:
                self.progress.emit("  [ERROR] Failed to load JSON files. Aborting.")
                self.finished.emit()
                return
            failed = len(files) - len(data)
            self.progress.emit(f"      > Loaded : {len(data):>5} / {len(files)}")
            if failed:
                self.progress.emit(f"      > Failed : {failed:>5}  (see load_log.txt)")

            # ── Step 3: analyze ───────────────────────────────────────────────
            self.progress.emit("[3/4] Analyzing files...")

            def analysis_cb(current, total):
                self.progress_update.emit("Analyzing files", current, total)

            results_df = self.controller._analyze_all_files(data, progress_callback=analysis_cb)
            passed = int(results_df['cond_final'].sum()) if 'cond_final' in results_df.columns else 0
            self.progress.emit(f"      > Pass : {passed:>5} / {len(results_df)}")
            self.progress.emit(f"      > Fail : {len(results_df) - passed:>5} / {len(results_df)}")
            self.controller._generate_summary_statistics()
            self.controller._save_results(results_df, self.controller.results_folder)
            self.progress.emit("      > Results saved to analysis_results.xlsx")

            # ── Step 4: visualizations ────────────────────────────────────────
            if self.generate_plots:
                self.progress.emit("[4/4] Creating signal visualizations (parallel)...")

                def viz_cb(current, total):
                    self.progress_update.emit("Creating visualizations", current, total)

                self.controller._create_sample_visualizations(
                    data, files, self.controller.visualizations_folder, progress_callback=viz_cb
                )
                self.progress.emit(f"      > {len(files)} plots saved")
            else:
                self.progress.emit("[4/4] Signal visualizations skipped.")
            self.finished.emit()
        except Exception as e:
            self.error.emit(str(e))
            self.finished.emit()

class CuffLDT(QMainWindow):
    def __init__(self, parent=None):
        """
        Initializes the Cuff Leakage Detection Tool (CuffLDT) application.
        This constructor sets up the main window and initializes the user interface.
        """
        super().__init__(parent)
        self.setWindowTitle("Cuff Leakage Detection Tool")
        self.setGeometry(100, 100, 600, 260)
        self.data_path = None
        self.results_folder = None
        self.visualizations_folder = None
        self._init_ui()

    def _init_ui(self):
        """
        Initializes the user interface for the Cuff Leakage Detection Tool.
        This method sets up the necessary components for user interaction.
        """

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        self.layout = QVBoxLayout()
        central_widget.setLayout(self.layout)

        # Read-only, scrollable log box (acts like a small terminal)
        class ReadOnlyLog(QPlainTextEdit):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self.setReadOnly(True)

            # keep a setText API so existing code that calls setText works
            def setText(self, text):
                self.setPlainText(text)

            def appendText(self, text):
                self.appendPlainText(text)

        self.status_label = ReadOnlyLog()
        self.status_label.setPlainText(
            "=" * 50 + "\n"
            + "  Cuff Leakage Detection Tool\n"
            + "=" * 50 + "\n"
            + "  Select a folder containing cuff run JSON\n"
            + "  files, then press Run Analysis.\n"
        )
        self.status_label.setFont(QFont("Courier", 10))
        self.status_label.setFixedHeight(120)
        self.layout.addWidget(self.status_label)

        # Progress tracking widgets (hidden until analysis starts)
        self.progress_step_label = QLabel("")
        self.progress_step_label.setFont(QFont("Courier", 9))
        self.progress_step_label.setVisible(False)
        self.layout.addWidget(self.progress_step_label)

        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(False)
        self.layout.addWidget(self.progress_bar)

        # Add Functionality buttons
        view_buttons = QHBoxLayout()
        self.layout.addLayout(view_buttons)
        self.select_folder_button = QPushButton("Select Folder")
        self.run_analysis_button = QPushButton("Run Analysis")
        self.open_results_button = QPushButton("Open Results")
        self.generate_plots_checkbox = QCheckBox("Generate plots")
        self.generate_plots_checkbox.setChecked(True)
        self.select_folder_button.clicked.connect(self.select_folder)
        self.run_analysis_button.clicked.connect(self.run_analysis)
        self.open_results_button.clicked.connect(self.open_results)
        view_buttons.addWidget(self.select_folder_button)
        view_buttons.addWidget(self.run_analysis_button)
        view_buttons.addWidget(self.open_results_button)
        view_buttons.addWidget(self.generate_plots_checkbox)
        self.select_folder_button.setEnabled(True)
        self.run_analysis_button.setEnabled(False)
        self.open_results_button.setEnabled(False)

    # =========================================================================
    # UI Methods
    # =========================================================================

    def select_folder(self):
        """Handles the folder selection process for the Cuff Leakage Detection Tool."""

        # Open file dialog to select folder containing cuff runs
        self.data_path = QFileDialog.getExistingDirectory(
            self,
            "Select Folder Containing Cuff Runs",
            DEFAULT_GEOMETRY_PATH,
        )

        # run checks on the selected folder
        if not self.data_path:
            QMessageBox.warning(self, "No Folder Selected", "Please select a folder to proceed.")
            self.status_label.appendText("No folder selected. Please select a folder to proceed.")
            return
        elif not os.path.isdir(self.data_path):
            QMessageBox.warning(self, "Invalid Folder", "The selected path is not a valid folder. Please select a valid folder.")
            self.status_label.appendText("Invalid folder selected. Please select a valid folder.")
            return
        elif not os.listdir(self.data_path):
            QMessageBox.warning(self, "Empty Folder", "The selected folder is empty. Please select a folder containing cuff runs.")
            self.status_label.appendText("Empty folder selected. Please select a folder containing cuff runs.")
            return
        elif not any(file.endswith('.json') for file in os.listdir(self.data_path)):
            QMessageBox.warning(self, "No JSON Files", "The selected folder does not contain any JSON files. Please select a folder containing cuff runs.")
            self.status_label.appendText("No JSON files found in the selected folder. Please select a folder containing cuff runs.")
            return
        
        # update the status label with the selected folder path and report some basic info
        files_present = os.listdir(self.data_path)
        num_json_files = sum(file.endswith('.json') for file in files_present)
        self.status_label.appendText(
            "\n" + "-" * 50 + "\n"
            f"[FOLDER] {self.data_path}\n"
            f"         JSON files found: {num_json_files}\n"
            + "-" * 50
        )
        
        # enable the run analysis button now that we have a valid folder selected
        self.run_analysis_button.setEnabled(True)
        
    def run_analysis(self):
        """Initiates the analysis process for the selected folder containing cuff runs."""
        # disable UI controls during analysis
        self.run_analysis_button.setEnabled(False)
        self.select_folder_button.setEnabled(False)
        self.generate_plots_checkbox.setEnabled(False)
        self.status_label.appendText(
            "\n" + "=" * 50 + "\n"
            "  Starting analysis...\n"
            + "=" * 50
        )

        self._thread = QThread()
        self._worker = AnalysisWorker(self, generate_plots=self.generate_plots_checkbox.isChecked())
        self._worker.moveToThread(self._thread)

        self._thread.started.connect(self._worker.run)
        self._worker.progress.connect(self.status_label.appendText)  # safe: signal handled in main thread
        self._worker.progress_update.connect(self._on_progress_update)
        self._worker.error.connect(lambda msg: QMessageBox.critical(self, "Analysis Error", msg))
        self._worker.finished.connect(self._on_analysis_finished)

        self.progress_bar.setValue(0)
        self.progress_bar.setVisible(True)
        self.progress_step_label.setVisible(True)

        self._worker.finished.connect(self._thread.quit)
        self._worker.finished.connect(self._worker.deleteLater)
        self._thread.finished.connect(self._thread.deleteLater)

        self._thread.start()

    def open_results(self):
        """Opens the results folder in the system's file explorer."""
        if not self.results_folder or not os.path.exists(self.results_folder):
            return
        
        if sys.platform == "darwin":
            subprocess.Popen(["open", self.results_folder])       # macOS
        elif sys.platform.startswith("win"):
            os.startfile(self.results_folder)                     # Windows
        else:
            subprocess.Popen(["xdg-open", self.results_folder])  # Linux

    # =========================================================================
    # Analysis Methods 
    # =========================================================================

    def _create_results_folder(self):
        """Create a results folder and visualizations subfolder within the data folder."""
        base_folder = os.path.dirname(self.data_path.rstrip('/\\'))
        self.results_folder = os.path.join(base_folder, "analysis_results")
        self.visualizations_folder = os.path.join(self.results_folder, "visualizations")
        os.makedirs(self.results_folder, exist_ok=True)
        os.makedirs(self.visualizations_folder, exist_ok=True)
        
    def _load_json_files(self, progress_callback=None):
        """Load all JSON files from the selected folder."""
        if not self.data_path:
            print("No folder selected. Exiting.")
            return None, None

        # Get all JSON files in the folder
        data_files = sorted([file for file in os.listdir(self.data_path) if file.endswith(".json")])

        if not data_files:
            return None, None

        # Load all JSON files
        data = {}
        failed_files = []
        total = len(data_files)
        for i, file in enumerate(data_files):
            if progress_callback:
                progress_callback(i + 1, total)
            try:
                with open(os.path.join(self.data_path, file), 'r') as f:
                    data[file] = json.load(f)
            except Exception as e:
                print(f"Error loading {file}: {e}")
                failed_files.append((file, str(e)))

        # summarize results
        total_files = len(data_files)
        loaded_count = len(data)
        failed_count = len(failed_files)

        # write a load log the user can inspect 
        log_path = os.path.join(self.results_folder, "load_log.txt") if self.results_folder else None
        if log_path:
            with open(log_path, 'w') as log_file:
                log_file.write(f"JSON load run: {datetime.now().isoformat()}\n")
                log_file.write(f"Data folder: {self.data_path}\n")
                log_file.write(f"Total JSON files found: {total_files}\n")
                log_file.write(f"Successfully loaded: {loaded_count}\n")
                log_file.write(f"Failed to load: {failed_count}\n\n")
                if failed_files:
                    log_file.write("Failed files and errors:\n")
                    for file, error in failed_files:
                        log_file.write(f"{file}: {error}\n")
        
        return data, data_files

    def _create_sample_visualizations(self, data, data_files, visualizations_folder, progress_callback=None):
        """Create signal visualizations for all JSON files using parallel worker processes."""
        sample_visualizations_folder = os.path.join(visualizations_folder, "signal_visualizations")
        os.makedirs(sample_visualizations_folder, exist_ok=True)

        args_list = [(data[f], f, sample_visualizations_folder) for f in data_files if f in data]
        total = len(args_list)

        with ProcessPoolExecutor(max_workers=os.cpu_count(),
                                 initializer=_init_plot_worker) as pool:
            futures = {pool.submit(_plot_worker, args): args[1] for args in args_list}
            for i, future in enumerate(as_completed(futures)):
                if progress_callback:
                    progress_callback(i + 1, total)
                try:
                    future.result()
                except Exception as e:
                    print(f"Visualization error: {e}")

    def _analyze_all_files(self, data, progress_callback=None):
        """Performs analysis on all loaded JSON files to detect cuff leaks."""
        results = []
        total = len(data)
        for i, (file_name, sample_data) in enumerate(data.items()):
            if progress_callback:
                progress_callback(i + 1, total)
            try:
                # partition the name to extract relevant info
                file_name = file_name.strip('.json')
                dev_names, run_name = file_name.split('-')
                cuff_id, ekg_id = self._parse_device_name(dev_names)

                metrics = self._analyze_single_file(sample_data)
                result = {
                    'file_name': file_name,
                    'cuff_id': cuff_id,
                    'ekg_id': ekg_id,
                    'run_name': run_name,
                    **metrics
                }
                results.append(result)
                
            except Exception as e:
                print(f"Error processing {file_name}: {e}")
                result = {
                    'file_name': file_name,
                    'cuff_id': None,
                    'ekg_id': None,
                    'run_name': None,
                    'cond_final': False,
                    'error': str(e)
                }
                results.append(result)
        
        return pd.DataFrame(results)

    def _generate_summary_statistics(self):
        """Generates summary statistics based on the analysis results."""
        pass

    def _save_results(self, results_df, results_folder):
        """Save results to CSV and Excel files with styling in Excel."""

        def highlight_failed_rows(row):
            """Highlight entire row red if condition is False."""
            if row['Final Pass/Fail'] == False:
                return ['background-color: #ffcccc'] * len(row)
            else:
                return [''] * len(row)

        # Create a copy with renamed columns for better readability
        column_rename_map = {
            'file_name': 'File Name',
            'cuff_id': 'Cuff ID',
            'ekg_id': 'EKG ID',
            'run_name': 'Run',
            'cond1': 'Cond 1 (6500 < Mean < 11000)',
            'cond1_st': 'Cond 1 (Mean < 11000)',
            'cond1_lt': 'Cond 1 (Mean > 6500)',
            'cond2': 'Cond 2 (Max < 14000)',
            'cond3': 'Cond 3 (Min > 2000)',
            'cond4': 'Cond 4 (Std < 1000)',
            'cond_final': 'Final Pass/Fail',
            'max_settled_value': 'Max Settled Value',
            'min_settled_value': 'Min Settled Value',
            'mean_settled_value': 'Mean Settled Value',
            'std_settled_value': 'Std Settled Value',
        }
        
        # Rename columns for display
        display_df = results_df.rename(columns=column_rename_map)

        # round to 0 decimal places columns related to settled values
        settled_value_cols = ['Max Settled Value', 'Min Settled Value', 'Mean Settled Value', 'Std Settled Value']
        display_df[settled_value_cols] = display_df[settled_value_cols].round(0)
        
        # Apply styling to highlight failed rows
        styled_df = display_df.style.apply(highlight_failed_rows, axis=1)
        styled_df = styled_df.set_properties(**{'text-align': 'center'})

        # create another sheet that only contains failed files
        failed_df = display_df[display_df['Final Pass/Fail'] == False]
        failed_styled_df = failed_df.style.apply(highlight_failed_rows, axis=1)
        failed_styled_df = failed_styled_df.set_properties(**{'text-align': 'center'})

        # Save to excel file
        excel_file = os.path.join(results_folder, "analysis_results.xlsx")
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # first save the styled dataframe to the first sheet
            styled_df.to_excel(writer, sheet_name='All Results', index=False)

            # then save the failed dataframe to the second sheet
            failed_styled_df.to_excel(writer, sheet_name='Failed Files', index=False)

        # format the excel sheet to have the columns auto-sized
        wb = load_workbook(excel_file)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2
        wb.save(excel_file)

        # styled_df.to_excel(excel_file, index=False, engine='openpyxl')
        print(f"Styled results saved to Excel: {excel_file}")
    # =========================================================================
    # SIGNAL ANALYSIS FUNCTIONS
    # =========================================================================

    def _extract_cuff_data(self, sample_data):
        """Extract cuff data from a sample JSON file."""
        tester_info = sample_data['tester_info']
        cuff_data = tester_info['cuff_data']
        cuff_values = np.array(cuff_data['cuff_values'])
        time_values = np.array(cuff_data['time'])
        return cuff_values, time_values

    def _process_ssbp_hold_signal(self, sample_data, cuff_values):
        """Process the sSBP hold signal to extract the relevant portion."""
        ssbp_hold = sample_data['hold_ssbp_cuff']
        length = len(ssbp_hold)
        ssbp_hold_data = cuff_values[-length:]

        # Compute last drop to cutoff signal
        gap = len(ssbp_hold_data) - np.argmax(np.abs(np.diff(ssbp_hold_data)))
        ssbp_hold_data = cuff_values[-(length+gap):-gap]
        
        return ssbp_hold_data

    def _analyze_single_file(self,sample_data):
        """Analyze a single JSON file and return metrics."""
        # Extract cuff data
        cuff_values, time_values = self._extract_cuff_data(sample_data)
        
        # Process sSBP hold signal
        ssbp_hold_data = self._process_ssbp_hold_signal(sample_data, cuff_values)

        # Cutoff the signal after the drop from the max value ~95% volts
        max_value = np.max(ssbp_hold_data)
        threshold = 0.95 * max_value
        cutoff_index = np.where(ssbp_hold_data > threshold)[0][-1] + 1
        ssbp_hold_data = ssbp_hold_data[cutoff_index:]

        # Look only at the signal that has settled (last 10000 samples)
        if len(ssbp_hold_data) > 10000:
            settled_signal = ssbp_hold_data[-10000:]
        else:
            settled_signal = ssbp_hold_data

        # compute the mean, max, min, and std of the settled signal
        mean_settled = np.mean(settled_signal)
        max_settled = np.max(settled_signal)
        min_settled = np.min(settled_signal)
        std_settled = np.std(settled_signal)

        # condition 1 - mean of settled signal < 11000 & > 6500
        cond1_st = mean_settled < 11000
        cond1_lt = mean_settled > 6500
        cond1 = cond1_st and cond1_lt

        # condition 2 - Max of settled signal < 14000
        cond2 = max_settled < 14000

        # condition 3 - Min of settled signal > 2000
        cond3 = min_settled > 2000

        # condition 4 - Std of settled signal < 1000
        cond4 = std_settled < 1000

        # Final condition
        cond = cond1 and cond2 and cond3 and cond4

        return {
            'max_settled_value': max_settled,
            'min_settled_value': min_settled,
            'mean_settled_value': mean_settled,
            'std_settled_value': std_settled,
            'cond1': cond1,
            'cond1_st': cond1_st,
            'cond1_lt': cond1_lt,
            'cond2': cond2,
            'cond3': cond3,
            'cond4': cond4,
            'cond_final': cond,
        }
    
    # =========================================================================
    # UTILITY METHODS
    # =========================================================================

    def _on_progress_update(self, label, current, total):
        pct = int(current / total * 100) if total > 0 else 0
        self.progress_bar.setValue(pct)
        self.progress_step_label.setText(f"{label}: {current}/{total}  ({pct}%)")

    def _on_analysis_finished(self):
        self.status_label.appendText(
            "=" * 50 + "\n"
            "  Analysis complete!\n"
            f"  Results: {self.results_folder}\n"
            + "=" * 50
        )
        self.progress_bar.setValue(100)
        self.progress_step_label.setText("Done.")
        self.select_folder_button.setEnabled(True)
        self.generate_plots_checkbox.setEnabled(True)
        self.run_analysis_button.setEnabled(True)
        self.open_results_button.setEnabled(True)

    def _parse_device_name(self, dev_name, cuff_prefix='CAA', ekg_prefix='PAA'):
        """
        Parse device name into cuff_id and ekg_id components.
        
        Args:
            dev_name: Device name string (e.g., 'CAA041PAA046')
            cuff_prefix: Prefix for cuff ID (default: 'CAA')
            ekg_prefix: Prefix for EKG ID (default: 'PAA')
        
        Returns:
            tuple: (cuff_id, ekg_id) or (None, None) if parsing fails
            
        Examples:
            >>> parse_device_name('CAA041PAA046')
            ('CAA041', 'PAA046')
            >>> parse_device_name('CAA1PAA2')
            ('CAA1', 'PAA2')
            >>> parse_device_name('CAA1234PAA56789')
            ('CAA1234', 'PAA56789')
        """
        # Create regex pattern: prefix followed by one or more digits
        pattern = f'^({cuff_prefix}\\d+)({ekg_prefix}\\d+)$'
        
        match = re.match(pattern, dev_name)
        if match:
            cuff_id = match.group(1)
            ekg_id = match.group(2)
            return cuff_id, ekg_id
        else:
            return None, None


if __name__ == "__main__":
    import multiprocessing
    multiprocessing.freeze_support()
    app = QApplication(sys.argv)
    window = CuffLDT()
    window.show()
    sys.exit(app.exec_())