import os
import sys
import json
import re
import tkinter as tk
from tkinter import filedialog, messagebox
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from tqdm import tqdm
import matplotlib
matplotlib.use("Agg")
plt.ioff()
import matplotlib.style as mplstyle
mplstyle.use('fast')
from openpyxl import load_workbook
import traceback
from datetime import datetime

# ============================================================================
# DATA LOADING AND FOLDER UTILITIES
# ============================================================================

def parse_device_name(dev_name, cuff_prefix='CAA', ekg_prefix='PAA'):
    """
    Parse device name into cuff_id and ekg_id components.
    
    Args:
        dev_name: Device name string (e.g., 'CAA041PAA046')
        cuff_prefix: Prefix for cuff ID (default: 'CAA')
        ekg_prefix: Prefix for EKG ID (default: 'PAA')
    
    Returns:
        tuple: (cuff_id, ekg_id) or (None, None) if parsing fails
        
    Examples:
        >>> parse_device_name('CAA041PAA046')
        ('CAA041', 'PAA046')
        >>> parse_device_name('CAA1PAA2')
        ('CAA1', 'PAA2')
        >>> parse_device_name('CAA1234PAA56789')
        ('CAA1234', 'PAA56789')
    """
    # Create regex pattern: prefix followed by one or more digits
    pattern = f'^({cuff_prefix}\\d+)({ekg_prefix}\\d+)$'
    
    match = re.match(pattern, dev_name)
    if match:
        cuff_id = match.group(1)
        ekg_id = match.group(2)
        return cuff_id, ekg_id
    else:
        return None, None


def _select_data_folder():
    """Open a dialog to select the folder containing JSON files."""
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    folder_path = filedialog.askdirectory(
        title="Select folder containing JSON files",
        initialdir=os.path.expanduser("~")
    )
    
    root.destroy()
    return folder_path


def _load_json_files(data_folder, save_folder=None):
    """Load all JSON files from the selected folder."""
    if not data_folder:
        print("No folder selected. Exiting.")
        return None, None
    
    # Get all JSON files in the folder
    data_files = sorted([file for file in os.listdir(data_folder) if file.endswith(".json")])
    
    if not data_files:
        messagebox.showerror("Error", "No JSON files found in the selected folder.")
        return None, None
    
    print(f"Number of data files in the folder: {len(data_files)}")
    
    # Load all JSON files
    data = {}
    failed_files = []
    for file in tqdm(data_files, total=len(data_files), desc="Loading JSON files"):
        try:
            with open(os.path.join(data_folder, file), 'r') as f:
                data[file] = json.load(f)
        except Exception as e:
            print(f"Error loading {file}: {e}")
            failed_files.append((file, str(e)))
    
    # summarize results
    total_files = len(data_files)
    loaded_count = len(data)
    failed_count = len(failed_files)

    # write a load log the user can inspect 
    log_path = os.path.join(save_folder, "load_log.txt") if save_folder else None
    if log_path:
        with open(log_path, 'w') as log_file:
            log_file.write(f"JSON load run: {datetime.now().isoformat()}\n")
            log_file.write(f"Data folder: {data_folder}\n")
            log_file.write(f"Total JSON files found: {total_files}\n")
            log_file.write(f"Successfully loaded: {loaded_count}\n")
            log_file.write(f"Failed to load: {failed_count}\n\n")
            if failed_files:
                log_file.write("Failed files and errors:\n")
                for file, error in failed_files:
                    log_file.write(f"{file}: {error}\n")
    
    return data, data_files


def _create_results_folder(data_folder):
    """Create a results folder and visualizations subfolder within the data folder."""
    base_folder = os.path.dirname(data_folder.rstrip('/\\'))
    results_folder = os.path.join(base_folder, "analysis_results")
    visualizations_folder = os.path.join(results_folder, "visualizations")
    os.makedirs(results_folder, exist_ok=True)
    os.makedirs(visualizations_folder, exist_ok=True)
    return results_folder, visualizations_folder


# ============================================================================
# SIGNAL ANALYSIS FUNCTIONS
# ============================================================================

def _extract_cuff_data(sample_data):
    """Extract cuff data from a sample JSON file."""
    tester_info = sample_data['tester_info']
    cuff_data = tester_info['cuff_data']
    cuff_values = np.array(cuff_data['cuff_values'])
    time_values = np.array(cuff_data['time'])
    return cuff_values, time_values


def _process_ssbp_hold_signal(sample_data, cuff_values):
    """Process the sSBP hold signal to extract the relevant portion."""
    ssbp_hold = sample_data['hold_ssbp_cuff']
    length = len(ssbp_hold)
    ssbp_hold_data = cuff_values[-length:]

    # Compute last drop to cutoff signal
    gap = len(ssbp_hold_data) - np.argmax(np.abs(np.diff(ssbp_hold_data)))
    ssbp_hold_data = cuff_values[-(length+gap):-gap]
    
    return ssbp_hold_data


def _analyze_single_file(sample_data):
    """Analyze a single JSON file and return metrics."""
    # Extract cuff data
    cuff_values, time_values = _extract_cuff_data(sample_data)
    
    # Process sSBP hold signal
    ssbp_hold_data = _process_ssbp_hold_signal(sample_data, cuff_values)

    # Cutoff the signal after the drop from the max value ~95% volts
    max_value = np.max(ssbp_hold_data)
    threshold = 0.95 * max_value
    cutoff_index = np.where(ssbp_hold_data > threshold)[0][-1] + 1
    ssbp_hold_data = ssbp_hold_data[cutoff_index:]

    # Look only at the signal that has settled (last 10000 samples)
    if len(ssbp_hold_data) > 10000:
        settled_signal = ssbp_hold_data[-10000:]
    else:
        settled_signal = ssbp_hold_data

    # compute the mean, max, min, and std of the settled signal
    mean_settled = np.mean(settled_signal)
    max_settled = np.max(settled_signal)
    min_settled = np.min(settled_signal)
    std_settled = np.std(settled_signal)

    # condition 1 - mean of settled signal < 11000 & > 6500
    cond1_st = mean_settled < 11000
    cond1_lt = mean_settled > 6500
    cond1 = cond1_st and cond1_lt

    # condition 2 - Max of settled signal < 14000
    cond2 = max_settled < 14000

    # condition 3 - Min of settled signal > 2000
    cond3 = min_settled > 2000

    # condition 4 - Std of settled signal < 1000
    cond4 = std_settled < 1000

    # Final condition
    cond = cond1 and cond2 and cond3 and cond4

    return {
        'max_settled_value': max_settled,
        'min_settled_value': min_settled,
        'mean_settled_value': mean_settled,
        'std_settled_value': std_settled,
        'cond1': cond1,
        'cond1_st': cond1_st,
        'cond1_lt': cond1_lt,
        'cond2': cond2,
        'cond3': cond3,
        'cond4': cond4,
        'cond_final': cond,
    }


def analyze_all_files(data, data_files):
    """Analyze all JSON files and return results as a DataFrame."""
    results = []
    
    for file_name, sample_data in tqdm(data.items(), total=len(data), desc="Analyzing files"):
        try:
            # partition the name to extract relevant info
            file_name = file_name.strip('.json')
            dev_names, run_name = file_name.split('-')
            cuff_id, ekg_id = parse_device_name(dev_names)

            metrics = _analyze_single_file(sample_data)
            result = {
                'file_name': file_name,
                'cuff_id': cuff_id,
                'ekg_id': ekg_id,
                'run_name': run_name,
                **metrics
            }
            results.append(result)
            
        except Exception as e:
            print(f"Error processing {file_name}: {e}")
            result = {
                'file_name': file_name,
                'cuff_id': None,
                'ekg_id': None,
                'run_name': None,
                'cond_final': False,
                'error': str(e)
            }
            results.append(result)
    
    return pd.DataFrame(results)


# ============================================================================
# PLOTTING AND VISUALIZATION FUNCTIONS
# ============================================================================

def _plot_sample_visualization(sample_data, file_name, visualizations_folder):
    """Create visualization for a sample file and save to visualizations folder."""
    # Extract and process data
    cuff_values, time_values = _extract_cuff_data(sample_data)
    ssbp_hold_data = _process_ssbp_hold_signal(sample_data, cuff_values)

    # Generate visualization
    fig, ax = plt.subplots(1, 2, figsize=(15, 6), gridspec_kw={'width_ratios': [2, 1]}, sharey=True)
    
    # Plot full signal
    ax[0].plot(time_values, cuff_values)
    ax[0].grid(axis='y')
    ax[0].set_title(f'Full Cuff Signal - {file_name}')
    ax[0].set_xlabel('Time (n)')
    ax[0].set_ylabel('Signal Amplitude')
    
    # Plot sSBP hold signal
    ax[1].plot(ssbp_hold_data, label='Signal')
    ax[1].axhline(np.mean(ssbp_hold_data[-10000:]), color='k', linestyle='-', label='Mean of Settled Signal')
    ax[1].grid(axis='y')
    ax[1].set_title('sSBP Hold Signal')
    ax[1].set_xlabel('Time (n)')
    ax[1].legend()
    ax[1].set_yticks(np.arange(0, 17000, 2000))
    
    plt.tight_layout()
    
    # Save the plot
    plot_filename = f"{os.path.splitext(file_name)[0]}.png"
    plot_path = os.path.join(visualizations_folder, plot_filename)
    plt.savefig(plot_path)
    plt.close(fig)

def create_sample_visualizations(data, data_files, visualizations_folder):
    """Create sample visualizations for all JSON files."""
    print(f"\nCreating sample visualizations for all {len(data_files)} files...")
    
    # Create the sample visualizations save folder
    sample_visualizations_folder = os.path.join(visualizations_folder, "signal_visualizations")
    os.makedirs(sample_visualizations_folder, exist_ok=True)
    
    for file_name in tqdm(data_files, desc="Creating sample visualizations"):
        try:
            sample_data = data[file_name]
            _plot_sample_visualization(sample_data, file_name, sample_visualizations_folder)
        except Exception as e:
            print(f"Error creating visualization for {file_name}: {e}")
    
    print("Sample visualizations completed!")

def create_summary_dashboard(results_df, visualizations_folder):
    """Create and save a summary dashboard with analysis statistics."""
    fig, axes = plt.subplots(2, 2, figsize=(12, 10))
    
    # Threshold compliance
    axes[0, 0].bar(['Cond1', 'Cond2', 'Cond3', 'Cond4', 'Final Cond'], 
                   [results_df['cond1'].sum(), results_df['cond2'].sum(), results_df['cond3'].sum(), results_df['cond4'].sum(), results_df['cond_final'].sum()])
    axes[0, 0].set_title('Threshold Compliance')
    axes[0, 0].set_ylabel('Number of Files')

    # for each column, add text at the base of the bar
    for i, v in enumerate([results_df['cond1'].sum(), results_df['cond2'].sum(), results_df['cond3'].sum(), results_df['cond4'].sum(), results_df['cond_final'].sum()]):
        axes[0, 0].text(i, 20, str(v), color='black', ha='center', va='bottom', bbox=dict(facecolor='white', alpha=0.5, edgecolor='black'))
    
    # Max values distribution
    axes[0, 1].hist(results_df['max_settled_value'].dropna(), bins=20, alpha=0.7)
    axes[0, 1].set_title('Distribution of Max Values')
    axes[0, 1].set_xlabel('Max Value')
    axes[0, 1].set_ylabel('Frequency')
    axes[0, 1].axvline(x=14000, color='r', linestyle='--', label='14000 threshold')
    axes[0, 1].legend()
    text = (f'Total Files: {len(results_df)}\n'
            f'Num >14000: {(results_df["max_settled_value"] > 14000).sum()}')
    axes[0, 1].text(0.95, 0.95, text, transform=axes[0, 1].transAxes,
                    verticalalignment='top', horizontalalignment='right')

    # Min settled values distribution
    axes[1, 0].hist(results_df['min_settled_value'].dropna(), bins=20, alpha=0.7)
    axes[1, 0].set_title('Distribution of Min Settled Values')
    axes[1, 0].set_xlabel('Min Settled Value')
    axes[1, 0].set_ylabel('Frequency')
    axes[1, 0].axvline(x=2000, color='r', linestyle='--', label='2000 threshold')
    axes[1, 0].legend()
    text = (f'Total Files: {len(results_df)}\n'
            f'Num <2000: {(results_df["min_settled_value"] < 2000).sum()}')
    axes[1, 0].text(0.95, 0.95, text, transform=axes[1, 0].transAxes,
                    verticalalignment='top', horizontalalignment='right')
    
    # Mean settled values distribution
    axes[1, 1].hist(results_df['mean_settled_value'].dropna(), bins=20, alpha=0.7)
    axes[1, 1].set_title('Distribution of Mean Settled Values')
    axes[1, 1].set_xlabel('Mean Settled Value')
    axes[1, 1].set_ylabel('Frequency')
    axes[1, 1].axvline(x=6500, color='r', linestyle='--', label='6500 threshold')
    axes[1, 1].axvline(x=11000, color='g', linestyle='--', label='11000 threshold')
    axes[1, 1].legend()

    # Add statistics text
    text = (f'Total Files: {len(results_df)}\n'
            f'Num <6500: {(results_df["mean_settled_value"] < 6500).sum()}\n'
            f'Num >11000: {(results_df["mean_settled_value"] > 11000).sum()}')
    axes[1, 1].text(0.95, 0.95, text, transform=axes[1, 1].transAxes,
                    verticalalignment='top', horizontalalignment='right')
    
    plt.tight_layout()
    
    # Save summary dashboard
    summary_plot_path = os.path.join(visualizations_folder, "summary_dashboard.png")
    plt.savefig(summary_plot_path, bbox_inches='tight')
    plt.close(fig)
    
    print(f"Summary dashboard saved to: {summary_plot_path}")


# ============================================================================
# RESULTS HANDLING
# ============================================================================

def save_results(results_df, results_folder):
    """Save results to CSV and Excel files with styling in Excel."""

    def highlight_failed_rows(row):
        """Highlight entire row red if condition is False."""
        if row['Final Pass/Fail'] == False:
            return ['background-color: #ffcccc'] * len(row)
        else:
            return [''] * len(row)

    # Create a copy with renamed columns for better readability
    column_rename_map = {
        'file_name': 'File Name',
        'cuff_id': 'Cuff ID',
        'ekg_id': 'EKG ID',
        'cond1': 'Condition 1 (Mean < 11000 & > 6500)',
        'cond2': 'Condition 2 (Max < 14000)',
        'cond3': 'Condition 3 (Min > 2000)',
        'cond4': 'Condition 4 (Std < 1000)',
        'cond_final': 'Final Pass/Fail',
        'max_settled_value': 'Max Settled Value',
        'min_settled_value': 'Min Settled Value',
        'mean_settled_value': 'Mean Settled Value',
        'std_settled_value': 'Std Settled Value'
    }
    
    # Rename columns for display
    display_df = results_df.rename(columns=column_rename_map)

    # round to 0 decimal places columns related to settled values
    settled_value_cols = ['Max Settled Value', 'Min Settled Value', 'Mean Settled Value', 'Std Settled Value']
    display_df[settled_value_cols] = display_df[settled_value_cols].round(0)
    
    # Apply styling to highlight failed rows
    styled_df = display_df.style.apply(highlight_failed_rows, axis=1)
    styled_df = styled_df.set_properties(**{'text-align': 'center'})

    # create another sheet that only contains failed files
    failed_df = display_df[display_df['Final Pass/Fail'] == False]
    failed_styled_df = failed_df.style.apply(highlight_failed_rows, axis=1)
    failed_styled_df = failed_styled_df.set_properties(**{'text-align': 'center'})

    # Save to excel file
    excel_file = os.path.join(results_folder, "analysis_results.xlsx")
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        # first save the styled dataframe to the first sheet
        styled_df.to_excel(writer, sheet_name='All Results', index=False)

        # then save the failed dataframe to the second sheet
        failed_styled_df.to_excel(writer, sheet_name='Failed Files', index=False)

    # format the excel sheet to have the columns auto-sized
    wb = load_workbook(excel_file)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2
    wb.save(excel_file)

    # styled_df.to_excel(excel_file, index=False, engine='openpyxl')
    print(f"Styled results saved to Excel: {excel_file}")

def print_summary_statistics(results_df):
    """Print summary statistics from the analysis."""
    print(f"\nSummary Statistics:")
    print(f"Total files analyzed: {len(results_df)}")
    print(f"Files meeting condition 1: {results_df['cond1'].sum()}")
    print(f"Files meeting condition 2: {results_df['cond2'].sum()}")
    print(f"Files meeting condition 3: {results_df['cond3'].sum()}")
    print(f"Files meeting condition 4: {results_df['cond4'].sum()}")
    print(f"Files meeting both conditions: {results_df['cond_final'].sum()}")


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Main function to run the batch analysis."""
    print("JSON Batch Analysis Tool")
    print("=" * 40)
    
    # Step 1: Select folder containing JSON files
    data_folder = _select_data_folder()
    if not data_folder:
        print("No folder selected. Exiting.")
        return
    print(f"Selected folder: {data_folder}")
    
    # Step 2: Create results and visualizations folders
    results_folder, visualizations_folder = _create_results_folder(data_folder)
    print(f"Results will be saved to: {results_folder}")
    print(f"Visualizations will be saved to: {visualizations_folder}")

    # Step 3: Load JSON files
    data, data_files = _load_json_files(data_folder, results_folder)
    if data is None:
        return
    
    # # Step 4: Create sample visualizations for all files
    # create_sample_visualizations(data, data_files, visualizations_folder)
    
    # Step 5: Analyze all files
    print("\nAnalyzing all files...")
    results_df = analyze_all_files(data, data_files)
    
    # Step 6: Display and save results
    print("\nAnalysis Results:")
    print("=" * 50)
    print(results_df)
    
    # Print summary statistics
    print_summary_statistics(results_df)
    
    # Save results to CSV
    save_results(results_df, results_folder)
    
    # Create and save summary dashboard
    create_summary_dashboard(results_df, visualizations_folder)
    
    print(f"\nAnalysis complete!")
    print(f"Results saved to: {results_folder}")
    print(f"Visualizations saved to: {visualizations_folder}")


if __name__ == "__main__":
    main()
